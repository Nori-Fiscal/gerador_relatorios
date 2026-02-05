import io
import re
import zipfile
import datetime as dt
import logging
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string


# ----------------------------
# Logging
# ----------------------------
logger = logging.getLogger("relatorio")
logger.setLevel(logging.INFO)


# ----------------------------
# Config model
# ----------------------------
@dataclass(frozen=True)
class ConfigCol:
    id_ref: int
    caminho: str
    titulo: str
    tipo: str
    is_produto: bool
    ordem_coluna: int
    is_geral: bool


def parse_bool(v: str) -> bool:
    return str(v).strip().lower() == "true"


def read_config_text(config_text: str) -> Tuple[str, List[ConfigCol], List[ConfigCol]]:
    lines = [ln.strip() for ln in config_text.splitlines() if ln.strip()]
    if not lines:
        raise ValueError("Config vazio.")

    header_parts = lines[0].split("|")
    model_name = header_parts[1].strip() if len(header_parts) >= 2 else "MODELO"

    cols: List[ConfigCol] = []
    for ln in lines[1:]:
        parts = ln.split("|")
        if len(parts) != 9:
            continue

        try:
            id_ref = int(parts[0].strip())
        except ValueError:
            continue

        caminho = parts[1].strip()
        titulo = parts[2].strip()
        tipo = parts[4].strip()
        is_produto = parse_bool(parts[5].strip())
        try:
            ordem = int(parts[7].strip())
        except ValueError:
            ordem = 999999
        is_geral = parse_bool(parts[8].strip())

        cols.append(
            ConfigCol(
                id_ref=id_ref,
                caminho=caminho,
                titulo=titulo,
                tipo=tipo,
                is_produto=is_produto,
                ordem_coluna=ordem,
                is_geral=is_geral,
            )
        )

    cols_geral = sorted([c for c in cols if c.is_geral], key=lambda c: c.ordem_coluna)
    cols_prod = sorted([c for c in cols if c.is_produto], key=lambda c: c.ordem_coluna)
    return model_name, cols_geral, cols_prod


# ----------------------------
# XML helpers (ignore namespace)
# ----------------------------
def local_name(tag: str) -> str:
    if "}" in tag:
        return tag.split("}", 1)[1]
    return tag


def first_child_by_local(parent: ET.Element, child_local: str) -> Optional[ET.Element]:
    for ch in list(parent):
        if local_name(ch.tag) == child_local:
            return ch
    return None


def find_first_by_local(root: ET.Element, local: str) -> Optional[ET.Element]:
    for el in root.iter():
        if local_name(el.tag) == local:
            return el
    return None


def get_inf_nfe(root: ET.Element) -> Optional[ET.Element]:
    return find_first_by_local(root, "infNFe")


# ----------------------------
# ICMS helper (ICMS00 / ICMSSNxxx etc.)
# ----------------------------
def find_in_icms_variants(icms_node: ET.Element, field_local: str) -> Optional[ET.Element]:
    targets = ["CST", "CSOSN"] if field_local == "CST" else [field_local]
    for variant in list(icms_node):
        for t in targets:
            el = first_child_by_local(variant, t)
            if el is not None:
                return el
    return None


# ----------------------------
# Extraction logic
# ----------------------------
def _parse_attr_syntax(segment: str) -> Tuple[str, Optional[str]]:
    s = segment.strip()
    if "/@" in s:
        tag_part, attr = s.split("/@", 1)
        return tag_part.strip(), attr.strip()
    if "@" in s and not s.startswith("@"):
        tag_part, attr = s.split("@", 1)
        return tag_part.strip(), attr.strip()
    if s.startswith("@"):
        return "", s[1:].strip()
    return s, None


def _get_text(el: Optional[ET.Element]) -> Optional[str]:
    if el is None or el.text is None:
        return None
    val = el.text.strip()
    return val if val else None


def extract_raw_value(
    base_node: ET.Element,
    caminho: str,
    *,
    item_index_1based: Optional[int] = None,
) -> Optional[str]:
    path = (caminho or "").strip()
    if not path:
        return None

    if path == "nItem":
        if "nItem" in base_node.attrib:
            v = base_node.attrib.get("nItem")
            return v.strip() if v else None
        if item_index_1based is not None:
            return str(item_index_1based)
        return None

    segments = [seg.strip() for seg in path.split(">") if seg.strip()]
    cur: ET.Element = base_node

    for seg in segments:
        tag_part, attr = _parse_attr_syntax(seg)

        # '@Id'
        if tag_part == "" and attr:
            v = cur.attrib.get(attr)
            if not v:
                return None
            v = v.strip()
            if attr == "Id" and v.startswith("NFe"):
                v = v[3:]  # fica só a chave numérica
            return v or None

        # base já é infNFe e caminho começa com infNFe/@Id
        if attr and tag_part and local_name(cur.tag) == tag_part:
            v = cur.attrib.get(attr)
            if not v:
                return None
            v = v.strip()
            if attr == "Id" and v.startswith("NFe"):
                v = v[3:]
            return v or None

        # CNPJCPF
        if tag_part == "CNPJCPF":
            cnpj = first_child_by_local(cur, "CNPJ")
            cpf = first_child_by_local(cur, "CPF")
            return _get_text(cnpj) or _get_text(cpf)

        nxt = first_child_by_local(cur, tag_part) if tag_part else None

        # ICMS inteligente
        if nxt is None and local_name(cur.tag) == "ICMS":
            alt = find_in_icms_variants(cur, tag_part)
            if alt is not None:
                cur = alt
                if attr:
                    v = cur.attrib.get(attr)
                    if not v:
                        return None
                    v = v.strip()
                    if attr == "Id" and v.startswith("NFe"):
                        v = v[3:]
                    return v or None
                continue

        if nxt is None:
            return None

        cur = nxt

        if attr:
            v = cur.attrib.get(attr)
            if not v:
                return None
            v = v.strip()
            if attr == "Id" and v.startswith("NFe"):
                v = v[3:]
            return v or None

    return _get_text(cur)


def extract_value_with_fallback_produtos(
    det_node: ET.Element,
    infNFe_node: ET.Element,
    caminho: str,
    *,
    item_index_1based: int,
) -> Optional[str]:
    v = extract_raw_value(det_node, caminho, item_index_1based=item_index_1based)
    if v is not None:
        return v
    return extract_raw_value(infNFe_node, caminho, item_index_1based=item_index_1based)


# ----------------------------
# CPF/CNPJ formatting
# ----------------------------
def only_digits(s: str) -> str:
    return re.sub(r"\D+", "", s or "")


def format_cpf_cnpj(value: str) -> str:
    d = only_digits(value)
    if len(d) == 11:
        return f"{d[0:3]}.{d[3:6]}.{d[6:9]}-{d[9:11]}"
    if len(d) == 14:
        return f"{d[0:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:14]}"
    return value


# ----------------------------
# Converters / typing
# ----------------------------
def _looks_like_iso_datetime(s: str) -> bool:
    return "T" in s and len(s) >= 10


def convert_value(raw: Optional[str], tipo: str, titulo_coluna: str) -> Any:
    if raw is None:
        return None

    t = (tipo or "").strip().upper()
    col = (titulo_coluna or "").strip().upper()

    # Formatar CNPJ/CPF com máscara
    if "CNPJ/CPF" in col or col in ("CNPJ", "CPF"):
        return format_cpf_cnpj(raw)

    if t == "REAL":
        try:
            return float(raw.replace(",", "."))
        except Exception:
            return raw

    if t == "DATA":
        s = raw.strip()
        try:
            if s.endswith("Z"):
                s = s[:-1] + "+00:00"
            if _looks_like_iso_datetime(s):
                dtt = dt.datetime.fromisoformat(s)
                return dtt.date()
            return dt.date.fromisoformat(s[:10])
        except Exception:
            return raw

    if t == "ENTRADA/SAIDA":
        s = raw.strip()
        if s == "0":
            return "ENTRADA"
        if s == "1":
            return "SAIDA"
        return raw

    if t == "NUMERO":
        s = raw.strip()
        if not s.isdigit():
            return raw

        preserve_as_text = any(k in col for k in ["CNPJ", "CPF", "EAN", "CFOP", "NCM", "CHAVE", "SKU"])
        if preserve_as_text or s.startswith("0") or len(s) > 10:
            return s

        try:
            return int(s)
        except Exception:
            return s

    return raw


# ----------------------------
# Excel helpers
# ----------------------------
CURRENCY_BR = "_-[$R$-pt-BR] * #,##0.00_-"


def set_header_style(ws, header_row: int = 1) -> None:
    bold = Font(bold=True)
    for cell in ws[header_row]:
        cell.font = bold


def auto_adjust_width(ws, max_widths: Dict[int, int], padding: int = 2, max_col_width: int = 60) -> None:
    for col_idx, w in max_widths.items():
        width = min(max(w + padding, 10), max_col_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def apply_column_formats(ws, col_types: List[str], header_row: int = 1) -> None:
    start_row = header_row + 1
    for i, tipo in enumerate(col_types, start=1):
        t = (tipo or "").strip().upper()
        if t not in ("REAL", "DATA"):
            continue

        for row in ws.iter_rows(min_row=start_row, min_col=i, max_col=i, max_row=ws.max_row):
            c = row[0]
            if c.value is None:
                continue
            if t == "REAL" and isinstance(c.value, (int, float)):
                c.number_format = "0.00"
            elif t == "DATA" and isinstance(c.value, (dt.date, dt.datetime)):
                c.number_format = "dd/mm/yyyy"


def normalize_excel_formulas(wb: Workbook) -> int:
    """
    Blindagem: fórmulas no XML do Excel devem usar ',' (invariante).
    Se alguma célula tiver ';', o Excel pode reparar/remover as fórmulas ao abrir.
    """
    changed = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                v = cell.value
                if cell.data_type == "f" or (isinstance(v, str) and v.startswith("=")):
                    if isinstance(v, str) and ";" in v:
                        cell.value = v.replace(";", ",")
                        changed += 1
    return changed


# ----------------------------
# "Macros" em Python (openpyxl)
# ----------------------------
def _last_row_by_col(ws, col_letter: str) -> int:
    col_idx = column_index_from_string(col_letter)
    for r in range(ws.max_row, 1, -1):
        v = ws.cell(r, col_idx).value
        if v is not None and str(v).strip() != "":
            return r
    return 1


def _to_float_if_numeric(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "":
        return None
    s2 = s.replace(".", "").replace(",", ".") if ("," in s and "." in s) else s.replace(",", ".")
    try:
        return float(s2)
    except Exception:
        return None


def _parse_date_br(v):
    if v is None:
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    s = str(v).strip()
    if not s:
        return None
    s = s.replace("-", "/")
    try:
        return dt.date.fromisoformat(s[:10].replace("/", "-"))
    except Exception:
        pass
    try:
        d, m, y = s[:10].split("/")
        return dt.date(int(y), int(m), int(d))
    except Exception:
        return None


def ajustar_dados_geral(ws):
    ultima = _last_row_by_col(ws, "B")
    if ultima < 2:
        return

    cols_num = ["A", "B", "K", "M", "N", "O", "S"]
    for col in cols_num:
        cidx = column_index_from_string(col)
        for r in range(2, ultima + 1):
            cell = ws.cell(r, cidx)
            fv = _to_float_if_numeric(cell.value)
            if fv is not None:
                cell.value = fv

    cols_money = ["K", "M", "N", "O", "S"]
    for col in cols_money:
        cidx = column_index_from_string(col)
        for r in range(2, ultima + 1):
            cell = ws.cell(r, cidx)
            if cell.value is None or str(cell.value).strip() == "":
                cell.value = 0
            cell.number_format = CURRENCY_BR

    cL = column_index_from_string("L")
    for r in range(2, ultima + 1):
        cell = ws.cell(r, cL)
        dv = _parse_date_br(cell.value)
        if dv is not None:
            cell.value = dv
            cell.number_format = "dd/mm/yyyy"

    ws.insert_rows(1)

    # Fórmulas SEMPRE com vírgula
    ws["D1"].value = "=SUBTOTAL(2,A:A)"
    ws["E1"].value = "=SUBTOTAL(9,K:K)"
    ws["E1"].number_format = CURRENCY_BR


def unificar_produtos(ws):
    ultima = _last_row_by_col(ws, "B")
    if ultima < 2:
        return

    cols_to_number = ["B", "J", "R", "T", "U", "V", "W", "X", "Z",
                      "AB", "AC", "AD", "AI", "AJ", "AM", "AN", "AR",
                      "AS", "AV", "AW", "AX", "AY", "AZ", "BD", "BE"]
    for col in cols_to_number:
        cidx = column_index_from_string(col)
        for r in range(2, ultima + 1):
            cell = ws.cell(r, cidx)
            fv = _to_float_if_numeric(cell.value)
            if fv is not None:
                cell.value = fv

    cols_money = ["J", "R", "T", "U", "V", "W", "X", "Z",
                  "AB", "AC", "AD", "AI", "AJ", "AM", "AN", "AR",
                  "AS", "AV", "AW", "AX", "AY", "AZ", "BD", "BE"]
    for col in cols_money:
        cidx = column_index_from_string(col)
        for r in range(2, ultima + 1):
            cell = ws.cell(r, cidx)
            if cell.value is None or str(cell.value).strip() == "":
                cell.value = 0
            cell.number_format = CURRENCY_BR

    cK = column_index_from_string("K")
    for r in range(2, ultima + 1):
        cell = ws.cell(r, cK)
        dv = _parse_date_br(cell.value)
        if dv is not None:
            cell.value = dv
            cell.number_format = "dd/mm/yyyy"
        else:
            if cell.value is not None:
                cell.value = str(cell.value).replace("-", "/")

    ws.insert_rows(1)

    # Fórmula SEMPRE com vírgula
    ws["C1"].value = "=SUBTOTAL(2,B:B)"

    check_cols = ["AA", "AE", "BA", "AV", "AR", "AK", "AL"]
    for col in check_cols:
        tmp = _last_row_by_col(ws, col)
        if tmp > ultima:
            ultima = tmp
    if ultima < 3:
        return

    def _fix_percent(col_letter: str):
        cidx = column_index_from_string(col_letter)
        for r in range(2, ultima + 1):
            cell = ws.cell(r, cidx)
            s = "" if cell.value is None else str(cell.value).strip()
            if s == "":
                cell.value = 0
            else:
                s = s.replace(".", ",")
                try:
                    val = float(s.replace(",", "."))
                    if s.startswith("0"):
                        cell.value = val
                    else:
                        cell.value = val / 100.0
                except Exception:
                    pass
            cell.number_format = "0.00%"

    _fix_percent("AA")
    _fix_percent("AE")

    cBA = column_index_from_string("BA")
    for r in range(2, ultima + 1):
        cell = ws.cell(r, cBA)
        s = "" if cell.value is None else str(cell.value).strip()
        if s == "":
            cell.value = 0
        else:
            try:
                v = float(s.replace(".", "").replace(",", "."))
                cell.value = v / 1_000_000
            except Exception:
                cell.value = 0
        cell.number_format = "0.00%"

    cAK = column_index_from_string("AK")
    cAL = column_index_from_string("AL")
    for r in range(2, ultima + 1):
        src = ws.cell(r, cAL).value
        if src is None or str(src).strip() == "":
            src = ws.cell(r, cAK).value
        s = "" if src is None else str(src).strip().replace(".", ",")
        if s == "":
            ws.cell(r, cAK).value = 0
        else:
            try:
                ws.cell(r, cAK).value = float(s.replace(",", ".")) / 100.0
            except Exception:
                ws.cell(r, cAK).value = 0
        ws.cell(r, cAK).number_format = "0.00%"

    colPIS = cAK
    colCof = column_index_from_string("AT")

    ws.delete_cols(cAL)
    if colCof > cAL:
        colCof -= 1

    ws.cell(2, colPIS).value = "Aliquota PIS"

    colAZ = column_index_from_string("AZ")
    colDifal = colAZ + 1
    ws.insert_cols(colDifal)
    ws.cell(1, colDifal).value = "DIFAL %"

    for r in range(2, ultima + 1):
        ws.cell(r, colDifal).value = f'=IF(AZ{r}=0,"",AZ{r}-AA{r})'
        ws.cell(r, colDifal).number_format = "0.00%"

    ws.cell(2, colCof).value = "Aliquota Cofins"

    cAW = column_index_from_string("AW")
    cAR = column_index_from_string("AR")
    cAU = column_index_from_string("AU")
    cAQ = column_index_from_string("AQ")

    for r in range(3, ultima + 1):
        aw = _to_float_if_numeric(ws.cell(r, cAW).value)
        ar = _to_float_if_numeric(ws.cell(r, cAR).value)
        au = _to_float_if_numeric(ws.cell(r, cAU).value)
        aq = _to_float_if_numeric(ws.cell(r, cAQ).value)

        out_cell = ws.cell(r, colCof)
        if aw is not None and ar not in (None, 0):
            out_cell.value = aw / ar
        elif au is not None and aq not in (None, 0):
            out_cell.value = au / aq
        else:
            out_cell.value = 0
        out_cell.number_format = "0.0%"


# ----------------------------
# Report generation helpers
# ----------------------------
def parse_xml_bytes_verbose(xml_bytes: bytes) -> Tuple[Optional[ET.Element], Optional[str]]:
    """
    Retorna (infNFe, motivo_erro). Se infNFe vier None, motivo_erro vem preenchido.
    """
    try:
        root = ET.fromstring(xml_bytes)
    except Exception as e:
        return None, f"XML inválido (parse): {type(e).__name__}: {str(e)[:240]}"

    inf = get_inf_nfe(root)
    if inf is None:
        return None, "Tag <infNFe> não encontrada (arquivo não parece ser NF-e, ou está incompleto/evento/CT-e)."

    return inf, None


def build_rows_for_nf(
    infNFe: ET.Element,
    cfg_geral: List[ConfigCol],
    cfg_prod: List[ConfigCol],
) -> Tuple[List[Any], List[List[Any]]]:
    row_geral = []
    for col in cfg_geral:
        raw = extract_raw_value(infNFe, col.caminho, item_index_1based=None)
        row_geral.append(convert_value(raw, col.tipo, col.titulo))

    dets = [ch for ch in list(infNFe) if local_name(ch.tag) == "det"]
    rows_prod: List[List[Any]] = []
    for idx, det in enumerate(dets, start=1):
        row = []
        for col in cfg_prod:
            raw = extract_value_with_fallback_produtos(det, infNFe, col.caminho, item_index_1based=idx)
            row.append(convert_value(raw, col.tipo, col.titulo))
        rows_prod.append(row)

    return row_geral, rows_prod


# ----------------------------
# ZIP routing (NORMAL / CANCELADAS / ignore INUTILIZADA)
# ----------------------------
@dataclass(frozen=True)
class XmlEntry:
    name: str
    content: bytes
    bucket: str  # "NORMAL" ou "CANCELADAS"
    zip_path: str


def _path_parts(p: str) -> List[str]:
    p = (p or "").replace("\\", "/")
    return [x.strip() for x in p.split("/") if x.strip()]


def _is_inutilizada(parts: List[str]) -> bool:
    return any(x.upper() == "INUTILIZADA" for x in parts)


def _is_cancelada(parts: List[str]) -> bool:
    return any("CANCELADA" in x.upper() for x in parts)


# ----------------------------
# ERROS (NFs não processadas)
# ----------------------------
@dataclass(frozen=True)
class ProcessError:
    zip_path: str
    name: str
    bucket: str
    etapa: str
    motivo: str
    detalhe: str


def _safe_str(v: Any, limit: int = 320) -> str:
    s = "" if v is None else str(v)
    s = s.replace("\r", " ").replace("\n", " ").strip()
    return s[:limit]


def _finalize_header_and_freeze(ws, expected_first_header: str):
    """
    Se a macro inseriu a linha 1 (subtotal), o header vira linha 2.
    Caso contrário, o header segue na linha 1.
    """
    if ws.max_row >= 2 and ws.cell(2, 1).value == expected_first_header:
        set_header_style(ws, header_row=2)
        ws.freeze_panes = "A3"
    else:
        set_header_style(ws, header_row=1)
        ws.freeze_panes = "A2"


def generate_workbook(
    xml_entries: List[XmlEntry],
    config_text: str,
    progress_cb=None,
) -> Tuple[Workbook, Dict[str, Any]]:
    model_name, cfg_geral, cfg_prod = read_config_text(config_text)

    wb = Workbook()
    wb.remove(wb.active)

    # Abas normais
    ws_geral = wb.create_sheet("GERAL")
    ws_prod = wb.create_sheet("PRODUTOS")

    # Abas canceladas (só se existirem)
    has_canceladas = any(e.bucket == "CANCELADAS" for e in xml_entries)
    ws_geral_c = wb.create_sheet("GERAL - CANCELADAS") if has_canceladas else None
    ws_prod_c = wb.create_sheet("PRODUTOS - CANCELADAS") if has_canceladas else None

    headers_geral = [c.titulo for c in cfg_geral]
    headers_prod = [c.titulo for c in cfg_prod]

    # headers (normais)
    ws_geral.append(headers_geral)
    ws_prod.append(headers_prod)
    set_header_style(ws_geral, header_row=1)
    set_header_style(ws_prod, header_row=1)
    ws_geral.freeze_panes = "A2"
    ws_prod.freeze_panes = "A2"

    # headers (canceladas)
    if has_canceladas and ws_geral_c and ws_prod_c:
        ws_geral_c.append(headers_geral)
        ws_prod_c.append(headers_prod)
        set_header_style(ws_geral_c, header_row=1)
        set_header_style(ws_prod_c, header_row=1)
        ws_geral_c.freeze_panes = "A2"
        ws_prod_c.freeze_panes = "A2"

    # widths
    widths_geral: Dict[int, int] = {i: len(h) for i, h in enumerate(headers_geral, start=1)}
    widths_prod: Dict[int, int] = {i: len(h) for i, h in enumerate(headers_prod, start=1)}
    widths_geral_c: Dict[int, int] = {i: len(h) for i, h in enumerate(headers_geral, start=1)} if has_canceladas else {}
    widths_prod_c: Dict[int, int] = {i: len(h) for i, h in enumerate(headers_prod, start=1)} if has_canceladas else {}

    ok = 0
    items = 0
    ok_canceladas = 0
    items_canceladas = 0
    skipped = 0

    errors: List[ProcessError] = []

    total = len(xml_entries)

    for i, entry in enumerate(xml_entries, start=1):
        if progress_cb and (i == 1 or i == total or i % 25 == 0):
            progress_cb(i, total, entry.zip_path)

        infNFe, motivo = parse_xml_bytes_verbose(entry.content)
        if infNFe is None:
            skipped += 1
            errors.append(
                ProcessError(
                    zip_path=entry.zip_path,
                    name=entry.name,
                    bucket=entry.bucket,
                    etapa="LEITURA XML",
                    motivo=_safe_str(motivo),
                    detalhe="",
                )
            )
            continue

        try:
            row_geral, rows_prod = build_rows_for_nf(infNFe, cfg_geral, cfg_prod)
        except Exception as e:
            skipped += 1
            errors.append(
                ProcessError(
                    zip_path=entry.zip_path,
                    name=entry.name,
                    bucket=entry.bucket,
                    etapa="EXTRAÇÃO",
                    motivo=f"{type(e).__name__}",
                    detalhe=_safe_str(str(e)),
                )
            )
            continue

        if entry.bucket == "CANCELADAS" and has_canceladas and ws_geral_c and ws_prod_c:
            t_ws_geral = ws_geral_c
            t_ws_prod = ws_prod_c
            t_wg = widths_geral_c
            t_wp = widths_prod_c
            ok_canceladas += 1
            items_canceladas += len(rows_prod)
        else:
            t_ws_geral = ws_geral
            t_ws_prod = ws_prod
            t_wg = widths_geral
            t_wp = widths_prod
            ok += 1
            items += len(rows_prod)

        t_ws_geral.append(row_geral)
        for ci, v in enumerate(row_geral, start=1):
            s = "" if v is None else str(v)
            t_wg[ci] = max(t_wg.get(ci, 0), len(s))

        for r in rows_prod:
            t_ws_prod.append(r)
            for ci, v in enumerate(r, start=1):
                s = "" if v is None else str(v)
                t_wp[ci] = max(t_wp.get(ci, 0), len(s))

    # formatos e widths (normais)
    apply_column_formats(ws_geral, [c.tipo for c in cfg_geral], header_row=1)
    apply_column_formats(ws_prod, [c.tipo for c in cfg_prod], header_row=1)
    auto_adjust_width(ws_geral, widths_geral)
    auto_adjust_width(ws_prod, widths_prod)

    # formatos e widths (canceladas)
    if has_canceladas and ws_geral_c and ws_prod_c:
        apply_column_formats(ws_geral_c, [c.tipo for c in cfg_geral], header_row=1)
        apply_column_formats(ws_prod_c, [c.tipo for c in cfg_prod], header_row=1)
        auto_adjust_width(ws_geral_c, widths_geral_c)
        auto_adjust_width(ws_prod_c, widths_prod_c)

    if progress_cb:
        progress_cb(total, total, "Aplicando ajustes (macros)")

    # macros (normais)
    ajustar_dados_geral(ws_geral)
    unificar_produtos(ws_prod)
    _finalize_header_and_freeze(ws_geral, expected_first_header=headers_geral[0] if headers_geral else "")
    _finalize_header_and_freeze(ws_prod, expected_first_header=headers_prod[0] if headers_prod else "")

    # macros (canceladas)
    if has_canceladas and ws_geral_c and ws_prod_c:
        ajustar_dados_geral(ws_geral_c)
        unificar_produtos(ws_prod_c)
        _finalize_header_and_freeze(ws_geral_c, expected_first_header=headers_geral[0] if headers_geral else "")
        _finalize_header_and_freeze(ws_prod_c, expected_first_header=headers_prod[0] if headers_prod else "")

    # aba de erros (se houver)
    if errors:
        ws_err = wb.create_sheet("NAO_PROCESSADAS")
        headers_err = ["Arquivo (ZIP/Origem)", "Bucket", "Etapa", "Motivo", "Detalhe", "Nome do Arquivo"]
        ws_err.append(headers_err)
        set_header_style(ws_err, header_row=1)
        ws_err.freeze_panes = "A2"

        widths_err: Dict[int, int] = {i: len(h) for i, h in enumerate(headers_err, start=1)}
        for er in errors:
            row = [er.zip_path, er.bucket, er.etapa, er.motivo, er.detalhe, er.name]
            ws_err.append(row)
            for ci, v in enumerate(row, start=1):
                s = "" if v is None else str(v)
                widths_err[ci] = max(widths_err.get(ci, 0), len(s))

        auto_adjust_width(ws_err, widths_err, max_col_width=80)

    stats: Dict[str, Any] = {
        "model": model_name,
        "ok": ok,
        "items": items,
        "ok_canceladas": ok_canceladas,
        "items_canceladas": items_canceladas,
        "skipped": skipped,
        "errors_count": len(errors),
        "cols_geral": len(cfg_geral),
        "cols_prod": len(cfg_prod),
        "has_canceladas": has_canceladas,
    }
    return wb, stats


# ----------------------------
# Streamlit UI
# ----------------------------
st.set_page_config(page_title="Relatório NF-e (XML → Excel)", layout="wide")
st.title("Relatório NF-e (XML → Excel)")
st.caption(
    "Envie vários XMLs ou um ZIP com XMLs. "
    "No ZIP: ignora pastas INUTILIZADA; pastas com CANCELADA vão para abas separadas. "
    "Se algum XML falhar, a aba 'NAO_PROCESSADAS' mostrará o motivo."
)

col1, col2 = st.columns(2)

with col1:
    uploaded_zip = st.file_uploader("ZIP com XMLs (recomendado para muitos arquivos)", type=["zip"])
    uploaded_xmls = st.file_uploader("Ou envie vários XMLs", type=["xml"], accept_multiple_files=True)

with col2:
    config_upload = st.file_uploader("Config (baserelatorio.txt) (opcional)", type=["txt"])
    default_config_path = st.text_input("Ou usar config do repositório (caminho)", value="baserelatorio.txt")
    out_name = st.text_input("Nome do arquivo de saída", value="relatorio.xlsx")


def load_config_text() -> str:
    if config_upload is not None:
        return config_upload.getvalue().decode("utf-8", errors="replace")
    p = Path(default_config_path)
    if not p.exists():
        raise FileNotFoundError(f"Config não encontrado no repositório: {p}. Faça upload do baserelatorio.txt.")
    return p.read_text(encoding="utf-8", errors="replace")


def collect_xml_entries() -> List["XmlEntry"]:
    entries: List[XmlEntry] = []

    if uploaded_zip is not None:
        zdata = uploaded_zip.getvalue()
        with zipfile.ZipFile(io.BytesIO(zdata), "r") as zf:
            for info in zf.infolist():
                if info.is_dir():
                    continue
                if not info.filename.lower().endswith(".xml"):
                    continue

                parts = _path_parts(info.filename)
                if _is_inutilizada(parts):
                    continue

                bucket = "CANCELADAS" if _is_cancelada(parts) else "NORMAL"
                entries.append(
                    XmlEntry(
                        name=Path(info.filename).name,
                        content=zf.read(info),
                        bucket=bucket,
                        zip_path=info.filename,
                    )
                )
        return entries

    if uploaded_xmls:
        for f in uploaded_xmls:
            entries.append(
                XmlEntry(
                    name=f.name,
                    content=f.getvalue(),
                    bucket="NORMAL",
                    zip_path=f.name,
                )
            )
        return entries

    return entries


btn = st.button("Gerar relatório", type="primary")

if btn:
    try:
        xml_entries = collect_xml_entries()
        if not xml_entries:
            st.error("Envie um ZIP com XMLs ou selecione vários XMLs.")
            st.stop()

        n_normal = sum(1 for e in xml_entries if e.bucket == "NORMAL")
        n_cancel = sum(1 for e in xml_entries if e.bucket == "CANCELADAS")
        st.caption(f"Arquivos considerados: NORMAL={n_normal} | CANCELADAS={n_cancel} | (INUTILIZADA ignorado)")

        config_text = load_config_text()

        progress = st.progress(0, text="Preparando...")
        status = st.empty()

        def progress_cb(done: int, total: int, current_name: str):
            pct = int((done / total) * 100) if total else 0
            progress.progress(pct, text=f"Processando: {done}/{total} ({pct}%)")
            status.write(f"Arquivo atual: `{current_name}`")

        with st.spinner("Gerando relatório..."):
            wb, stats = generate_workbook(xml_entries, config_text, progress_cb=progress_cb)
            changed = normalize_excel_formulas(wb)

            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)

        progress.progress(100, text="Concluído!")
        st.success("Relatório gerado com sucesso!")

        st.write(
            f"**Modelo:** {stats['model']}  \n"
            f"**NF-es válidas (NORMAL):** {stats['ok']} | **Itens (NORMAL):** {stats['items']}  \n"
            f"**NF-es válidas (CANCELADAS):** {stats['ok_canceladas']} | **Itens (CANCELADAS):** {stats['items_canceladas']}  \n"
            f"**Puladas (inválidas):** {stats['skipped']}  \n"
            f"**Com motivo registrado (aba NAO_PROCESSADAS):** {stats['errors_count']}  \n"
            f"**Fórmulas normalizadas (troca ';'→','):** {changed}"
        )

        if stats["errors_count"] > 0:
            st.info("Abra o Excel e veja a aba **NAO_PROCESSADAS** para entender o motivo de cada XML não processado.")

        st.download_button(
            label="Baixar Excel",
            data=bio.getvalue(),
            file_name=out_name if out_name.lower().endswith(".xlsx") else (out_name + ".xlsx"),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_excel_relatorio",
        )

    except Exception as e:
        st.exception(e)
